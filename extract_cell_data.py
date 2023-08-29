def extract_cell_data(filepath, sheet_name, cell_address_list):
    """
    Extracts data from specific cells within a given worksheet in an Excel spreadsheet and returns it as a dictionary.

    The dictionary keys correspond to the cell addresses and the values correspond to the data within those cells.

    Parameters:
    - filepath (str): The complete path to the spreadsheet file, including its extension.
    - sheet_name (str): The name of the worksheet within the spreadsheet from which to extract data.
    - cell_address_list (List[str]): A list containing the addresses of the cells to extract (e.g., ['A1', 'B2', 'C3']).

    Returns:
    - dict: A dictionary where keys are the cell addresses and values are the cell contents.

    Example:
    >>> cell_data = extract_cell_data('path_to_your_spreadsheet.xlsx', 'Sheet1', ['A1', 'B2', 'C3'])
    >>> print(cell_data)
    # Output could be: {'A1': 'Name', 'B2': 32, 'C3': 'Some Data'}

    Note:
    - This function uses the `openpyxl` library and is applicable to Excel files (.xlsx).
    - Raises an error if the specified file, worksheet, or cell addresses do not exist.
    """  
    from openpyxl import load_workbook
    
    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    
    data = {}
    for cell_address in cell_address_list:
        data[cell_address] = ws[cell_address].value
    
    wb.close()
    
    return data